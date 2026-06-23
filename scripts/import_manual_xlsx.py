from __future__ import annotations

import argparse
import os
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_HOME = Path(os.getenv("MINSAENG100_HOME", Path(__file__).resolve().parents[1]))
DB_PATH = Path(os.getenv("MINSAENG100_DB", APP_HOME / "data" / "minsaeng100.sqlite"))

CREDIT_SHEET = "신용보증공급"
POLICY_SHEET = "정책자금"
HEADER_ROW = 3
DATA_START_ROW = 4

CREDIT_COLUMNS = {
    "기준월": "base_month",
    "보증공급액(원)": "guarantee_supply_amount_krw",
    "보증공급건수(건)": "guarantee_supply_count",
    "보증잔액(원)": "guarantee_balance_krw",
    "자료제공기관": "source_org",
    "원본파일명": "source_file_name",
    "입력자": "input_user",
    "입력일": "input_date",
    "비고": "note",
}

POLICY_COLUMNS = {
    "기준월": "base_month",
    "사업명": "program_name",
    "총계획금액(원)": "total_plan_amount_krw",
    "누계지원액(원)": "cumulative_support_amount_krw",
    "누계지원건수(건)": "cumulative_support_count",
    "집행률(%)": "execution_rate_pct",
    "자료제공기관": "source_org",
    "출처URL": "source_url",
    "원본파일명": "source_file_name",
    "입력자": "input_user",
    "입력일": "input_date",
    "비고": "note",
}


def normalize_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def normalize_month(value: Any) -> str | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        value = str(int(value))
    value = str(value).replace(".", "-").replace("/", "-").strip()
    if len(value) == 6 and value.isdigit():
        return f"{value[:4]}-{value[4:]}"
    return value[:7]


def to_int(value: Any) -> int | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace(",", "")
    return int(float(value))


def to_float(value: Any) -> float | None:
    value = normalize_cell(value)
    if value is None:
        return None
    if isinstance(value, str):
        value = value.replace(",", "")
    return float(value)


def to_text(value: Any) -> str | None:
    value = normalize_cell(value)
    return None if value is None else str(value)


def clean_note(value: Any) -> str | None:
    note = to_text(value)
    if note and note.startswith("예시 행"):
        return None
    return note


def header_map(ws, columns: dict[str, str]) -> dict[str, int]:
    found = {}
    for cell in ws[HEADER_ROW]:
        if cell.value in columns:
            found[columns[cell.value]] = cell.column
    missing = [target for target in columns.values() if target not in found]
    if missing:
        raise ValueError(f"{ws.title} 시트 필수 컬럼 누락: {', '.join(missing)}")
    return found


def iter_sheet_rows(ws, columns: dict[str, str]):
    mapping = header_map(ws, columns)
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        row = {name: ws.cell(row=row_idx, column=col_idx).value for name, col_idx in mapping.items()}
        if not any(normalize_cell(value) is not None for value in row.values()):
            continue
        yield row


def choose_policy_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        base_month = normalize_month(row.get("base_month"))
        if not base_month:
            continue
        row["_base_month"] = base_month
        grouped.setdefault(base_month, []).append(row)

    selected_rows = []
    for base_month, month_rows in grouped.items():
        if len(month_rows) == 1:
            selected_rows.append(month_rows[0])
            continue
        total_rows = [
            row
            for row in month_rows
            if to_text(row.get("program_name")) == "소상공인 특별자금"
        ]
        if total_rows:
            selected_rows.append(total_rows[0])
            continue
        raise ValueError(
            f"{base_month} 정책자금 행이 여러 건입니다. "
            "DB에는 월별 총괄 1건만 저장하므로 '소상공인 특별자금' 총괄 행을 지정해야 합니다."
        )
    return selected_rows


def import_credit(conn: sqlite3.Connection, ws) -> int:
    written = 0
    for row in iter_sheet_rows(ws, CREDIT_COLUMNS):
        base_month = normalize_month(row.get("base_month"))
        if not base_month:
            continue
        conn.execute(
            """
            INSERT INTO manual_credit_guarantee_monthly (
                base_month, guarantee_supply_amount_krw, guarantee_supply_count,
                guarantee_balance_krw, source_org, source_file_name,
                input_user, input_date, note, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(base_month) DO UPDATE SET
                guarantee_supply_amount_krw=excluded.guarantee_supply_amount_krw,
                guarantee_supply_count=excluded.guarantee_supply_count,
                guarantee_balance_krw=excluded.guarantee_balance_krw,
                source_org=excluded.source_org,
                source_file_name=excluded.source_file_name,
                input_user=excluded.input_user,
                input_date=excluded.input_date,
                note=excluded.note,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                base_month,
                to_int(row.get("guarantee_supply_amount_krw")),
                to_int(row.get("guarantee_supply_count")),
                to_int(row.get("guarantee_balance_krw")),
                to_text(row.get("source_org")) or "부산신용보증재단",
                to_text(row.get("source_file_name")),
                to_text(row.get("input_user")) or "관리자",
                to_text(row.get("input_date")) or date.today().isoformat(),
                clean_note(row.get("note")),
            ),
        )
        written += 1
    return written


def import_policy(conn: sqlite3.Connection, ws) -> int:
    written = 0
    for row in choose_policy_rows(list(iter_sheet_rows(ws, POLICY_COLUMNS))):
        base_month = row.get("_base_month") or normalize_month(row.get("base_month"))
        if not base_month:
            continue
        total = to_int(row.get("total_plan_amount_krw")) or 0
        amount = to_int(row.get("cumulative_support_amount_krw"))
        execution_rate = to_float(row.get("execution_rate_pct"))
        if execution_rate is None and amount is not None and total:
            execution_rate = round(amount / total * 100, 2)
        conn.execute(
            """
            INSERT INTO manual_policy_fund_monthly (
                base_month, program_name, total_plan_amount_krw,
                cumulative_support_amount_krw, cumulative_support_count,
                execution_rate_pct, source_org, source_url, source_file_name,
                input_user, input_date, note, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(base_month) DO UPDATE SET
                program_name=excluded.program_name,
                total_plan_amount_krw=excluded.total_plan_amount_krw,
                cumulative_support_amount_krw=excluded.cumulative_support_amount_krw,
                cumulative_support_count=excluded.cumulative_support_count,
                execution_rate_pct=excluded.execution_rate_pct,
                source_org=excluded.source_org,
                source_url=excluded.source_url,
                source_file_name=excluded.source_file_name,
                input_user=excluded.input_user,
                input_date=excluded.input_date,
                note=excluded.note,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                base_month,
                to_text(row.get("program_name")) or "소상공인 특별자금",
                total,
                amount,
                to_int(row.get("cumulative_support_count")),
                execution_rate,
                "부산신용보증재단",
                to_text(row.get("source_url")),
                to_text(row.get("source_file_name")),
                to_text(row.get("input_user")) or "관리자",
                to_text(row.get("input_date")) or date.today().isoformat(),
                clean_note(row.get("note")),
            ),
        )
        written += 1
    return written


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    args = parser.parse_args()

    if not DB_PATH.exists():
        raise SystemExit(f"DB not found: {DB_PATH}. Run scripts/init_db.py first.")

    xlsx_path = Path(args.xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    if CREDIT_SHEET not in wb.sheetnames or POLICY_SHEET not in wb.sheetnames:
        raise SystemExit(f"필수 시트가 없습니다: {CREDIT_SHEET}, {POLICY_SHEET}")

    with sqlite3.connect(DB_PATH) as conn:
        credit_written = import_credit(conn, wb[CREDIT_SHEET])
        policy_written = import_policy(conn, wb[POLICY_SHEET])
        conn.execute(
            """
            INSERT INTO import_runs (source, source_ref, status, rows_written, finished_at, message)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
            """,
            (
                "manual_xlsx",
                str(xlsx_path),
                "success",
                credit_written + policy_written,
                f"신용보증 {credit_written}건, 정책자금 {policy_written}건 적재",
            ),
        )

    print(f"imported xlsx: credit={credit_written}, policy={policy_written}, path={xlsx_path}")


if __name__ == "__main__":
    main()
