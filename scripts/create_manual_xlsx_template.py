from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


APP_HOME = Path(__file__).resolve().parents[1]
OUT_PATH = APP_HOME / "data" / "manual" / "민생100일_정책금융_수동입력_서식.xlsx"

NAVY = "1F4E79"
BLUE = "0070C0"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREEN = "E2F0D9"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"
RED = "C00000"
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


CREDIT_HEADERS = [
    ("기준월", "base_month", "YYYY-MM 형식. 예: 2026-05"),
    ("보증공급액(원)", "guarantee_supply_amount_krw", "해당 월 보증공급 금액. 원 단위 정수 입력"),
    ("보증공급건수(건)", "guarantee_supply_count", "해당 월 보증공급 건수. 없으면 공란 가능"),
    ("보증잔액(원)", "guarantee_balance_krw", "월말 보증잔액. 없으면 공란 가능"),
    ("자료제공기관", "source_org", "기본값: 부산신용보증재단"),
    ("원본파일명", "source_file_name", "기관 제공 원본 파일명 또는 문서명"),
    ("입력자", "input_user", "입력 담당자명"),
    ("입력일", "input_date", "YYYY-MM-DD 형식"),
    ("비고", "note", "특이사항"),
]

POLICY_HEADERS = [
    ("기준월", "base_month", "YYYY-MM 형식. 예: 2026-05"),
    ("사업명", "program_name", "기본값: 소상공인 특별자금"),
    ("총계획금액(원)", "total_plan_amount_krw", "사업 총 계획금액. 기본값 592,500,000,000원"),
    ("누계지원액(원)", "cumulative_support_amount_krw", "기준월까지 누계 지원금액. 원 단위 정수 입력"),
    ("누계지원건수(건)", "cumulative_support_count", "기준월까지 누계 지원건수. 없으면 공란 가능"),
    ("집행률(%)", "execution_rate_pct", "누계지원액 / 총계획금액 * 100. 자동 계산"),
    ("자료제공기관", "source_org", "기본값: 부산일포유/부산광역시/부산신용보증재단"),
    ("출처URL", "source_url", "사업 안내 또는 자료 출처 URL"),
    ("원본파일명", "source_file_name", "원본 파일명 또는 문서명"),
    ("입력자", "input_user", "입력 담당자명"),
    ("입력일", "input_date", "YYYY-MM-DD 형식"),
    ("비고", "note", "특이사항"),
]


def style_guide(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "민생100일 정책금융 수동 입력서식"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:F1")

    rows = [
        ("사용 목적", "부산신용보증재단 월별 보증공급 규모와 소상공인 특별자금 누계 집행률을 대시보드 DB에 적재하기 위한 입력서식입니다."),
        ("작성 시트", "신용보증공급, 정책자금"),
        ("기준월 형식", "YYYY-MM 예: 2026-05"),
        ("금액 단위", "원 단위로 입력. 화면에서는 억원 단위로 변환됩니다."),
        ("집행률", "정책자금 시트의 집행률(%)은 누계지원액 / 총계획금액 * 100으로 자동 계산됩니다."),
        ("주의", "소상공인 특별자금은 전체 신용보증 공급규모의 세부 프로그램이므로 두 금액을 합산하지 않습니다."),
        ("DB 적재 명령", "python scripts/import_manual_xlsx.py data/manual/민생100일_정책금융_수동입력_서식.xlsx"),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        ws.cell(row=row_idx, column=1).font = Font(name="Arial", bold=True, color=NAVY)
        ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(row=row_idx, column=1).border = BORDER
        ws.cell(row=row_idx, column=2).border = BORDER
        ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 115


def setup_input_sheet(ws, title: str, headers: list[tuple[str, str, str]], table_name: str, rows: int = 36) -> int:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = "노란색 셀은 필수 입력 또는 주요 입력 항목입니다. 금액은 원 단위로 입력합니다."
    ws["A2"].font = Font(name="Arial", color=RED, bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for col_idx, (label, field, comment) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(f"{field}\n{comment}", "Codex")

    end_row = 3 + rows
    for row_idx in range(4, end_row + 1):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.font = Font(name="Arial", color=BLUE)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

    table_ref = f"A3:{get_column_letter(len(headers))}{end_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = [13, 22, 18, 20, 30, 30, 14, 14, 38, 18, 14, 38]
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1] if idx <= len(widths) else 18
    ws.row_dimensions[3].height = 34

    month_validation = DataValidation(type="textLength", operator="between", formula1="7", formula2="7", allow_blank=True)
    ws.add_data_validation(month_validation)
    month_validation.add(f"A4:A{end_row}")

    for col_idx, (label, _, _) in enumerate(headers, start=1):
        col = get_column_letter(col_idx)
        if "(원)" in label or "(건)" in label:
            for row_idx in range(4, end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        if "(%)" in label:
            for row_idx in range(4, end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.00"
        if label == "입력일":
            for row_idx in range(4, end_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"
    return end_row


def main() -> None:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "작성가이드"
    ws_credit = wb.create_sheet("신용보증공급")
    ws_policy = wb.create_sheet("정책자금")

    style_guide(ws_guide)
    setup_input_sheet(ws_credit, "부산신용보증재단 월별 보증공급 입력", CREDIT_HEADERS, "tbl_credit_input")
    policy_end = setup_input_sheet(ws_policy, "소상공인 특별자금 누계 집행 입력", POLICY_HEADERS, "tbl_policy_input")

    ws_credit["A4"] = "2026-05"
    ws_credit["E4"] = "부산신용보증재단"
    ws_credit["G4"] = "관리자"
    ws_credit["H4"] = "2026-06-10"
    ws_credit["I4"] = "예시 행. 실제 금액 입력 후 사용"

    ws_policy["A4"] = "2026-05"
    ws_policy["B4"] = "소상공인 특별자금"
    ws_policy["C4"] = 592_500_000_000
    ws_policy["G4"] = "부산일포유/부산광역시/부산신용보증재단"
    ws_policy["H4"] = "https://bsjob4u.co.kr/support/view.php?sprt_plc_seq=36"
    ws_policy["J4"] = "관리자"
    ws_policy["K4"] = "2026-06-10"
    ws_policy["L4"] = "예시 행. 누계지원액 입력 시 집행률 자동 계산"

    for row_idx in range(4, policy_end + 1):
        cell = ws_policy.cell(row=row_idx, column=6)
        cell.value = f'=IF(OR(C{row_idx}="",D{row_idx}="",C{row_idx}=0),"",ROUND(D{row_idx}/C{row_idx}*100,2))'
        cell.font = Font(name="Arial", color=BLACK)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        cell.protection = Protection(locked=False)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
